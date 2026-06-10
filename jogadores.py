"""
Coletor de estatísticas individuais de jogadores do SofaScore.
Reusa BrowserSession do coletor.py para evitar bloqueio 403.

Uso: python jogadores.py
"""

import difflib
import json
import os
import unicodedata
import time
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from coletor import BrowserSession

# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------
_CACHE_FILE = "cache_jogadores.json"
_cache: dict = {}


def _load_cache():
    global _cache
    if os.path.exists(_CACHE_FILE):
        with open(_CACHE_FILE, "r", encoding="utf-8") as f:
            _cache = json.load(f)
        print(f"[cache] {len(_cache)} estatística(s) de jogadores em cache.\n")


def _save_cache():
    with open(_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(_cache, f, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Mapeamento de estatísticas
# ---------------------------------------------------------------------------
_STAT_KEYS = [
    ("goals",                  "gols"),
    ("goalAssist",             "assistencias"),
    ("totalShots",             "finalizacoes"),
    ("onTargetScoringAttempt", "chutes_no_gol"),
    ("foulCommitted",          "faltas_cometidas"),
    ("wasFouled",              "faltas_sofridas"),
    ("yellowCard",             "cartao_amarelo"),
    ("redCard",                "cartao_vermelho"),
    ("minutesPlayed",          "minutos"),
    ("rating",                 "rating"),
]

_NUM_COLS = [
    ("gols",              "Gols"),
    ("assistencias",      "Assistências"),
    ("finalizacoes",      "Finalizações"),
    ("chutes_no_gol",     "Chutes no Gol"),
    ("faltas_cometidas",  "Faltas Cometidas"),
    ("faltas_sofridas",   "Faltas Sofridas"),
    ("cartao_amarelo",    "Cart. Amarelo"),
    ("cartao_vermelho",   "Cart. Vermelho"),
    ("minutos",           "Minutos"),
    ("rating",            "Rating"),
]

_METRICS = [
    ("Média",          lambda s: s.mean()),
    ("Desvio Padrão",  lambda s: s.std()),
    ("Mediana",        lambda s: s.median()),
    ("Mínimo",         lambda s: s.min()),
    ("Máximo",         lambda s: s.max()),
    ("Média - Desvio", lambda s: s.mean() - s.std()),
]

# ---------------------------------------------------------------------------
# Paleta de cores (espelho do coletor.py)
# ---------------------------------------------------------------------------
_C_HEADER  = "2E75B6"
_C_TEAM    = "1F4E79"
_C_METRIC  = "D6E4F0"
_C_WHITE   = "FFFFFF"
_C_ROW_ALT = "EBF3FB"
_C_RESULTADO = {"Vitória": "C6EFCE", "Empate": "FFEB9C", "Derrota": "FFC7CE"}


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 22)


# ---------------------------------------------------------------------------
# Funções de coleta
# ---------------------------------------------------------------------------

def fetch_team_players(session: BrowserSession, team_id: int) -> list[dict]:
    """Retorna lista de jogadores de um time."""
    try:
        data = session.get_json(f"/team/{team_id}/players")
        return [
            {
                "id":       p["player"]["id"],
                "name":     p["player"]["name"],
                "position": p["player"].get("position", "?"),
            }
            for p in data.get("players", [])
        ]
    except RuntimeError as e:
        print(f"  [ERRO] Jogadores do time {team_id}: {e}")
        return []


def fetch_player_last_games(
    session: BrowserSession,
    player_id: int,
    n_games: int = 5,
    team_id: int | None = None,
) -> list:
    """Busca os últimos N jogos finalizados de um jogador.

    Se team_id for informado, retorna apenas jogos onde esse time
    (normalmente a seleção) participou — filtrando jogos pelo clube.
    Busca mais páginas automaticamente até encontrar N jogos válidos.
    """
    all_finished = []
    page = 0
    max_pages = 20  # limite de segurança

    while len(all_finished) < n_games and page < max_pages:
        try:
            data   = session.get_json(f"/player/{player_id}/events/last/{page}")
            events = data.get("events", [])
        except RuntimeError as e:
            print(f"  [ERRO] Jogos do jogador {player_id} (p{page}): {e}")
            break

        if not events:
            break

        finished = [e for e in events if e.get("status", {}).get("type") == "finished"]

        if team_id is not None:
            finished = [
                e for e in finished
                if e.get("homeTeam", {}).get("id") == team_id
                or e.get("awayTeam", {}).get("id") == team_id
            ]

        all_finished.extend(finished)
        page += 1

        if len(data.get("events", [])) < 10:
            break

        time.sleep(0.3)

    all_finished.sort(key=lambda x: x.get("startTimestamp", 0), reverse=True)
    return all_finished[:n_games]


def fetch_player_event_stats(session: BrowserSession, event_id: int, player_id: int) -> dict:
    """Busca estatísticas de um jogador em uma partida específica (com cache)."""
    key = f"{event_id}_{player_id}"
    if key in _cache:
        return _cache[key]
    try:
        data = session.get_json(f"/event/{event_id}/player/{player_id}/statistics")
        stats = data.get("statistics", {})
        _cache[key] = stats
        _save_cache()
        return stats
    except RuntimeError as e:
        print(f"  [ERRO] Stats jogador {player_id} evento {event_id}: {e}")
        return {}


def collect_player_stats(
    session: BrowserSession,
    player_id: int,
    player_name: str,
    team_id: int,
    team_name: str,
    n_games: int = 5,
    selecao_only: bool = False,
) -> list:
    """Pipeline: busca jogos → stats por jogo → monta linhas para o DataFrame.

    selecao_only=True filtra apenas jogos onde team_id (a seleção) participou.
    """
    print(f">>> [{player_name}]  id={player_id}  ({team_name})"
          + ("  [apenas seleção]" if selecao_only else ""))

    games = fetch_player_last_games(
        session, player_id, n_games,
        team_id=team_id if selecao_only else None,
    )
    if not games:
        print("  Nenhum jogo encontrado.")
        return []

    rows = []
    for game in games:
        event_id   = game.get("id")
        home_team  = game.get("homeTeam", {})
        away_team  = game.get("awayTeam", {})
        home_name  = home_team.get("name", "?")
        away_name  = away_team.get("name", "?")
        home_score = game.get("homeScore", {}).get("current", 0)
        away_score = game.get("awayScore", {}).get("current", 0)
        ts         = game.get("startTimestamp", 0)
        date_str   = datetime.fromtimestamp(ts).strftime("%d/%m/%Y") if ts else "?"
        tournament = game.get("tournament", {}).get("name", "?")

        # Determina em qual lado o jogador estava
        if home_team.get("id") == team_id:
            side, opponent = "home", away_name
            gols_f, gols_s = home_score, away_score
        elif away_team.get("id") == team_id:
            side, opponent = "away", home_name
            gols_f, gols_s = away_score, home_score
        else:
            side, opponent = None, f"{home_name} x {away_name}"
            gols_f, gols_s = None, None

        if gols_f is not None:
            resultado = "Vitória" if gols_f > gols_s else ("Derrota" if gols_f < gols_s else "Empate")
            local     = "Casa" if side == "home" else "Fora"
        else:
            resultado = "?"
            local     = "?"

        print(f"  [{date_str}] {home_name} {home_score}x{away_score} {away_name}  (id={event_id})")

        api_stats = fetch_player_event_stats(session, event_id, player_id)

        row = {
            "jogador":        player_name,
            "time":           team_name,
            "data":           date_str,
            "competicao":     tournament,
            "adversario":     opponent,
            "local":          local,
            "placar":         f"{home_score}x{away_score}",
            "resultado":      resultado,
        }
        for api_key, col_key in _STAT_KEYS:
            row[col_key] = api_stats.get(api_key)

        rows.append(row)
        time.sleep(0.8)

    return rows


# ---------------------------------------------------------------------------
# Exportação Excel
# ---------------------------------------------------------------------------

def _format_data_sheet(ws):
    col_labels = {
        "jogador": "Jogador", "time": "Time", "data": "Data",
        "competicao": "Competição", "adversario": "Adversário",
        "local": "Local", "placar": "Placar", "resultado": "Resultado",
        **{col: lbl for col, lbl in _NUM_COLS},
    }
    for cell in ws[1]:
        cell.value     = col_labels.get(cell.value, cell.value)
        cell.font      = Font(bold=True, color=_C_WHITE)
        cell.fill      = _fill(_C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[1].height = 20

    res_col = next((c.column for c in ws[1] if c.value == "Resultado"), None)
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        resultado_val = ws.cell(i, res_col).value if res_col else None
        row_fill = _fill(_C_RESULTADO.get(resultado_val, _C_ROW_ALT if i % 2 == 0 else _C_WHITE))
        for cell in row:
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _append_resumo(ws, pdf: pd.DataFrame, start_row: int):
    present_cols = [(col, lbl) for col, lbl in _NUM_COLS if col in pdf.columns]
    row = start_row

    # Título
    ws.cell(row, 1).value     = "RESUMO ESTATÍSTICO"
    ws.cell(row, 1).font      = Font(bold=True, color=_C_WHITE, size=11)
    ws.cell(row, 1).fill      = _fill(_C_TEAM)
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(present_cols) + 1)
    ws.row_dimensions[row].height = 20
    row += 1

    # Cabeçalho de colunas
    ws.cell(row, 1).value     = "Métrica"
    ws.cell(row, 1).font      = Font(bold=True, color=_C_WHITE)
    ws.cell(row, 1).fill      = _fill(_C_HEADER)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border    = _thin_border()
    for j, (_, lbl) in enumerate(present_cols, 2):
        c           = ws.cell(row, j)
        c.value     = lbl
        c.font      = Font(bold=True, color=_C_WHITE)
        c.fill      = _fill(_C_HEADER)
        c.alignment = Alignment(horizontal="center")
        c.border    = _thin_border()
    row += 1

    # Linhas de métricas
    metrics_start = row
    for m_idx, (metric_name, func) in enumerate(_METRICS):
        row_fill        = _fill(_C_METRIC) if m_idx % 2 == 0 else _fill(_C_WHITE)
        lbl_cell        = ws.cell(row, 1)
        lbl_cell.value     = metric_name
        lbl_cell.font      = Font(bold=True)
        lbl_cell.fill      = row_fill
        lbl_cell.alignment = Alignment(horizontal="left")
        lbl_cell.border    = _thin_border()

        for j, (col, _) in enumerate(present_cols, 2):
            series = pd.to_numeric(pdf[col], errors="coerce").dropna()
            if len(series) >= 2:
                val = round(float(func(series)), 2)
            elif len(series) == 1:
                val = round(float(series.iloc[0]), 2)
            else:
                val = None
            c               = ws.cell(row, j)
            c.value         = val
            c.fill          = row_fill
            c.alignment     = Alignment(horizontal="center")
            c.border        = _thin_border()
            c.number_format = "0.00"
        row += 1

    # Formatação condicional por coluna
    for j in range(2, len(present_cols) + 2):
        col_letter = get_column_letter(j)
        ws.conditional_formatting.add(
            f"{col_letter}{metrics_start}:{col_letter}{row - 1}",
            ColorScaleRule(
                start_type="min",      start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max",        end_color="63BE7B",
            ),
        )


def export_to_excel(all_rows: list, filename: str = "jogadores_stats.xlsx"):
    if not all_rows:
        print("\nNenhum dado coletado para exportar.")
        return

    df = pd.DataFrame(all_rows)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Aba Todos
        df.to_excel(writer, sheet_name="Todos", index=False)
        _format_data_sheet(writer.sheets["Todos"])

        # Aba por jogador: dados + resumo
        for player in df["jogador"].unique():
            pdf        = df[df["jogador"] == player]
            sheet_name = str(player)[:31]
            pdf.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _format_data_sheet(ws)
            _append_resumo(ws, pdf, len(pdf) + 3)

    print(f"\nExportado para: {filename}")


# ---------------------------------------------------------------------------
# CLI interativo
# ---------------------------------------------------------------------------

_ALL_COPA_TEAMS: dict[int, str] = {
    4691: "Argélia",       4819: "Argentina",     4741: "Austrália",
    4718: "Áustria",       4717: "Bélgica",        4479: "Bósnia e Herzegovina",
    4748: "Brasil",        4753: "Cabo Verde",     4752: "Canadá",
    4820: "Colômbia",      4715: "Croácia",        55827: "Curaçao",
    4714: "Rep. Tcheca",   4768: "Costa do Marfim", 4823: "Congo RD",
    4757: "Equador",       4758: "Egito",          4713: "Inglaterra",
    4481: "França",        4711: "Alemanha",       4764: "Gana",
    7229: "Haiti",         4766: "Irã",            4767: "Iraque",
    4770: "Japão",         4771: "Jordânia",       4781: "México",
    4778: "Marrocos",      4705: "Países Baixos",  4784: "Nova Zelândia",
    4475: "Noruega",       5164: "Panamá",         4789: "Paraguai",
    4704: "Portugal",      4792: "Catar",          4834: "Arábia Saudita",
    4695: "Escócia",       4739: "Senegal",        4736: "África do Sul",
    4735: "Coreia do Sul", 4698: "Espanha",        4688: "Suécia",
    4699: "Suíça",         4729: "Tunísia",        4700: "Turquia",
    4724: "Estados Unidos", 4725: "Uruguai",       4723: "Uzbequistão",
}

_POS_LABEL = {"G": "GOL", "D": "DEF", "M": "MEI", "F": "ATA"}


def _normalize(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def _find(query: str, pool: dict) -> list:
    norm_q = _normalize(query)
    exact  = [(k, v) for k, v in pool.items() if _normalize(v) == norm_q]
    if exact:
        return exact
    prefix = [(k, v) for k, v in pool.items() if _normalize(v).startswith(norm_q)]
    if prefix:
        return prefix
    close  = difflib.get_close_matches(norm_q, [_normalize(v) for v in pool.values()], n=3, cutoff=0.5)
    return [(k, v) for k, v in pool.items() if _normalize(v) in close]


def _pick_team(session: BrowserSession) -> tuple[int, str] | None:
    """Mostra a lista de times e deixa o usuário escolher um."""
    sorted_teams = sorted(_ALL_COPA_TEAMS.items(), key=lambda x: x[1])
    col_w = max(len(n) for n in _ALL_COPA_TEAMS.values()) + 2
    print("\nTimes da Copa:")
    for i, (_, name) in enumerate(sorted_teams, 1):
        print(f"  {name:<{col_w}}", end="\n" if i % 3 == 0 else "")
    print("\n")

    while True:
        try:
            raw = input("Digite o time para ver o elenco (ou 'sair'): ").strip()
        except (EOFError, KeyboardInterrupt):
            return None
        if raw.lower() == "sair":
            return None
        matches = _find(raw, dict(sorted_teams))
        if not matches:
            print(f"  Time nao encontrado: '{raw}'\n")
            continue
        if len(matches) == 1:
            return matches[0]
        print("  Mais de um resultado:")
        for i, (_, name) in enumerate(matches, 1):
            print(f"    {i}. {name}")
        try:
            choice = input("  Numero: ").strip()
            if choice.isdigit() and 1 <= int(choice) <= len(matches):
                return matches[int(choice) - 1]
        except (EOFError, KeyboardInterrupt):
            return None


def select_players(session: BrowserSession) -> list[dict]:
    """Loop interativo de seleção de jogadores."""
    selected: list[dict] = []
    roster_cache: dict[int, list] = {}

    print("\n" + "=" * 56)
    print("    COPA DO MUNDO 2026 — ESTATÍSTICAS DE JOGADORES")
    print("=" * 56)

    while True:
        team_result = _pick_team(session)
        if team_result is None:
            break
        team_id, team_name = team_result

        if team_id not in roster_cache:
            print(f"  Buscando elenco de {team_name}...")
            roster_cache[team_id] = fetch_team_players(session, team_id)

        roster = roster_cache[team_id]
        roster_dict = {p["id"]: f"{p['name']} ({_POS_LABEL.get(p['position'], p['position'])})" for p in roster}

        # Mostra o elenco
        print(f"\n  {team_name} — {len(roster)} jogadores:")
        for p in sorted(roster, key=lambda x: x["position"] + x["name"]):
            pos = _POS_LABEL.get(p["position"], p["position"])
            print(f"    [{pos}] {p['name']}")
        print()
        print("  -> Digite o nome do jogador para adicionar.")
        print("  -> 'todos' para selecionar o elenco inteiro.")
        print("  -> 'outro' para escolher outro time.")
        print("  -> 'pronto' para finalizar.\n")

        while True:
            sel_names = [s["name"] for s in selected]
            sel_txt   = ", ".join(sel_names) if sel_names else "nenhum"
            try:
                raw = input(f"[Selecionados: {sel_txt}]\n> ").strip()
            except (EOFError, KeyboardInterrupt):
                print()
                return selected

            if not raw:
                continue

            cmd = raw.lower()

            if cmd == "pronto":
                return selected

            if cmd == "outro":
                break

            if cmd == "todos":
                for p in roster:
                    if not any(s["id"] == p["id"] for s in selected):
                        selected.append({"id": p["id"], "name": p["name"],
                                         "team_id": team_id, "team_name": team_name})
                print(f"  + {len(roster)} jogadores de {team_name} adicionados.\n")
                continue

            # Busca por nome dentro do elenco
            name_dict = {p["id"]: p["name"] for p in roster}
            matches   = _find(raw, name_dict)

            if not matches:
                print(f"  Jogador nao encontrado: '{raw}'\n")
                continue

            if len(matches) == 1:
                pid, pname = matches[0]
                if not any(s["id"] == pid for s in selected):
                    selected.append({"id": pid, "name": pname,
                                     "team_id": team_id, "team_name": team_name})
                    print(f"  + {pname} adicionado.\n")
                else:
                    print(f"  {pname} ja esta na lista.\n")
            else:
                print("  Mais de um resultado:")
                for i, (_, name) in enumerate(matches, 1):
                    print(f"    {i}. {name}")
                try:
                    choice = input("  Numero (Enter para cancelar): ").strip()
                    if choice.isdigit() and 1 <= int(choice) <= len(matches):
                        pid, pname = matches[int(choice) - 1]
                        if not any(s["id"] == pid for s in selected):
                            selected.append({"id": pid, "name": pname,
                                             "team_id": team_id, "team_name": team_name})
                            print(f"  + {pname} adicionado.\n")
                except (EOFError, KeyboardInterrupt):
                    print()

    return selected


def _ask_int(prompt: str, default: int) -> int:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
        return int(raw) if raw else default
    except (ValueError, EOFError, KeyboardInterrupt):
        return default


def _ask_str(prompt: str, default: str) -> str:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
        return raw if raw else default
    except (EOFError, KeyboardInterrupt):
        return default


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    _load_cache()
    session = BrowserSession()
    try:
        players = select_players(session)
        if not players:
            print("Nenhum jogador selecionado.")
        else:
            n_games = _ask_int("\nQuantos jogos recentes por jogador?", default=5)
            outfile = _ask_str("Nome do arquivo Excel?", default="jogadores_stats.xlsx")
            print()

            all_rows = []
            for p in players:
                rows = collect_player_stats(
                    session, p["id"], p["name"], p["team_id"], p["team_name"], n_games
                )
                all_rows.extend(rows)

            export_to_excel(all_rows, filename=outfile)
    finally:
        session.close()
