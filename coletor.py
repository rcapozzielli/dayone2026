"""
Coletor de estatísticas do SofaScore via endpoints JSON internos.
Usa Playwright para simular um browser real (necessário para evitar bloqueio 403).

Uso:
  python coletor.py
  (o script pergunta quais times, quantos jogos e o nome do arquivo)

Dependências:
  pip install playwright pandas openpyxl
  python -m playwright install chromium
"""

import difflib
import json
import os
import sys as _sys
import unicodedata
import time
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from playwright.sync_api import sync_playwright, Page

# ---------------------------------------------------------------------------
# Cache de estatísticas de eventos (evita rebuscar jogos já coletados)
# ---------------------------------------------------------------------------
_CACHE_FILE = "cache_stats.json"
_cache: dict = {}


def _load_cache():
    global _cache
    if os.path.exists(_CACHE_FILE):
        with open(_CACHE_FILE, "r", encoding="utf-8") as f:
            _cache = json.load(f)
        print(f"[cache] {len(_cache)} evento(s) em cache carregados.\n")


def _save_cache():
    with open(_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(_cache, f, ensure_ascii=False)

# ---------------------------------------------------------------------------
# Mapeamento de estatísticas que queremos extrair
# (grupo_keyword, stat_keyword, chave_de_saida)
# ---------------------------------------------------------------------------
STAT_MAP = [
    ("Shots",          "Total shots",     "finalizacoes"),
    ("Shots",          "Shots on target", "chutes_no_gol"),
    ("Match overview", "Corner kicks",    "escanteios"),
    ("Match overview", "Yellow cards",    "cartoes_amarelos"),
    ("Match overview", "Red cards",       "cartoes_vermelhos"),
]

BASE_URL = "https://api.sofascore.com/api/v1"
_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)


# ---------------------------------------------------------------------------
# Camada de HTTP via Playwright
# O browser fica aberto durante toda a coleta; cada chamada de API é
# executada como fetch() dentro do contexto do browser, evitando o bloqueio
# 403 por TLS-fingerprinting / challenge que o SofaScore impõe a requests.
# ---------------------------------------------------------------------------

class BrowserSession:
    """Wrapper fino sobre uma página Playwright para chamadas à API."""

    def __init__(self):
        self._pw = sync_playwright().start()
        _args = []
        if _sys.platform != "win32":
            _args += ["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]
        self._browser = self._pw.chromium.launch(headless=True, args=_args)
        self._ctx = self._browser.new_context(user_agent=_UA)
        self._page: Page = self._ctx.new_page()
        self._warm_up()

    def _warm_up(self):
        """Visita a página principal para estabelecer sessão/cookies."""
        print("Iniciando sessão no SofaScore...")
        self._page.goto("https://www.sofascore.com/", wait_until="networkidle", timeout=60_000)
        time.sleep(3)
        print("Sessão pronta.\n")

    def get_json(self, path: str) -> dict:
        """Executa fetch(path) dentro do browser e retorna o JSON parseado."""
        url = f"{BASE_URL}{path}"
        result = self._page.evaluate(
            """async (url) => {
                const resp = await fetch(url, {
                    headers: {
                        'Accept': 'application/json, text/plain, */*',
                        'Referer': 'https://www.sofascore.com/'
                    }
                });
                if (!resp.ok) return { __error: resp.status };
                return await resp.json();
            }""",
            url,
        )
        if isinstance(result, dict) and "__error" in result:
            raise RuntimeError(f"HTTP {result['__error']} para {url}")
        return result or {}

    def close(self):
        self._browser.close()
        self._pw.stop()


# ---------------------------------------------------------------------------
# FUNÇÕES PRINCIPAIS
# ---------------------------------------------------------------------------

def fetch_last_games(session: BrowserSession, team_id: int, n_games: int = 5) -> list:
    """Busca os últimos N jogos finalizados de um time, paginando se necessário."""
    all_finished = []
    page = 0

    while len(all_finished) < n_games:
        try:
            data = session.get_json(f"/team/{team_id}/events/last/{page}")
            events = data.get("events", [])
        except Exception as e:
            raise RuntimeError(f"Falha ao buscar jogos do time {team_id} (p{page}): {e}") from e

        if not events:
            break

        finished = [e for e in events if e.get("status", {}).get("type") == "finished"]
        all_finished.extend(finished)
        page += 1

        if len(events) < 10:
            break

        time.sleep(0.5)

    all_finished.sort(key=lambda x: x.get("startTimestamp", 0), reverse=True)
    return all_finished[:n_games]


def fetch_event_statistics(session: BrowserSession, event_id: int) -> dict:
    """Busca estatísticas de uma partida (usa cache se disponível)."""
    key = str(event_id)
    if key in _cache:
        return _cache[key]
    try:
        data = session.get_json(f"/event/{event_id}/statistics")
        _cache[key] = data
        _save_cache()
        return data
    except RuntimeError as e:
        print(f"  [ERRO] Estatísticas do evento {event_id}: {e}")
        return {}


def fetch_event_incidents(session: BrowserSession, event_id: int) -> list:
    """Busca incidentes de uma partida (gols, cartões, substituições)."""
    try:
        return session.get_json(f"/event/{event_id}/incidents").get("incidents", [])
    except RuntimeError as e:
        print(f"  [ERRO] Incidentes do evento {event_id}: {e}")
        return []


def extract_stats(statistics_data: dict, side: str) -> dict:
    """
    Extrai estatísticas relevantes para um lado ('home' ou 'away').
    Usa apenas o período 'ALL' (jogo completo).
    """
    result = {key: None for _, _, key in STAT_MAP}

    for period_block in statistics_data.get("statistics", []):
        if period_block.get("period") != "ALL":
            continue
        for group in period_block.get("groups", []):
            group_name = group.get("groupName", "")
            for item in group.get("statisticsItems", []):
                stat_name = item.get("name", "")
                for (g_kw, s_kw, out_key) in STAT_MAP:
                    if g_kw.lower() in group_name.lower() and s_kw.lower() in stat_name.lower():
                        result[out_key] = item.get(side)

    return result


def collect_team_stats(
    session: BrowserSession, team_id: int, team_name: str = None, n_games: int = 5
) -> list:
    """Pipeline completo: busca jogos → estatísticas → monta linhas para o DataFrame."""
    label = team_name or str(team_id)
    print(f">>> [{label}]  team_id={team_id}")

    games = fetch_last_games(session, team_id, n_games)
    if not games:
        print("  Nenhum jogo finalizado encontrado.")
        return []

    rows = []
    for game in games:
        event_id   = game.get("id")
        home_team  = game.get("homeTeam", {}).get("name", "?")
        away_team  = game.get("awayTeam", {}).get("name", "?")
        home_id    = game.get("homeTeam", {}).get("id")
        ts         = game.get("startTimestamp", 0)
        date_str   = datetime.fromtimestamp(ts).strftime("%d/%m/%Y") if ts else "?"
        tournament = game.get("tournament", {}).get("name", "?")

        home_score = game.get("homeScore", {}).get("current", 0)
        away_score = game.get("awayScore", {}).get("current", 0)

        side      = "home" if home_id == team_id else "away"
        opponent  = away_team if side == "home" else home_team
        gols_f    = home_score if side == "home" else away_score
        gols_s    = away_score if side == "home" else home_score

        if gols_f > gols_s:
            resultado = "Vitória"
        elif gols_f < gols_s:
            resultado = "Derrota"
        else:
            resultado = "Empate"

        print(f"  [{date_str}] {home_team} {home_score}x{away_score} {away_team}  (id={event_id})")

        stats_data   = fetch_event_statistics(session, event_id)
        extracted    = extract_stats(stats_data, side)
        adv_side     = "away" if side == "home" else "home"
        adv_extracted = {f"adv_{k}": v for k, v in extract_stats(stats_data, adv_side).items()}

        row = {
            "time":          label,
            "data":          date_str,
            "competicao":    tournament,
            "adversario":    opponent,
            "local":         "Casa" if side == "home" else "Fora",
            "placar":        f"{home_score}x{away_score}",
            "gols_feitos":   gols_f,
            "gols_sofridos": gols_s,
            "resultado":     resultado,
            **extracted,
            **adv_extracted,
        }
        rows.append(row)
        time.sleep(1)

    return rows


# Colunas numéricas e seus rótulos legíveis
_NUM_COLS = [
    ("gols_feitos",           "Gols Feitos"),
    ("gols_sofridos",         "Gols Sofridos"),
    ("finalizacoes",          "Finalizações"),
    ("chutes_no_gol",         "Chutes no Gol"),
    ("escanteios",            "Escanteios"),
    ("cartoes_amarelos",      "Cart. Amarelos"),
    ("cartoes_vermelhos",     "Cart. Vermelhos"),
    ("adv_finalizacoes",      "Adv. Finalizações"),
    ("adv_chutes_no_gol",     "Adv. Chutes no Gol"),
    ("adv_escanteios",        "Adv. Escanteios"),
    ("adv_cartoes_amarelos",  "Adv. Cart. Amarelos"),
    ("adv_cartoes_vermelhos", "Adv. Cart. Vermelhos"),
]

# Métricas da aba Resumo
_METRICS = [
    ("Média",          lambda s: s.mean()),
    ("Desvio Padrão",  lambda s: s.std()),
    ("Mediana",        lambda s: s.median()),
    ("Mínimo",         lambda s: s.min()),
    ("Máximo",         lambda s: s.max()),
    ("Média - Desvio", lambda s: s.mean() - s.std()),
    ("Mín - Desvio",   lambda s: s.min() - s.std()),
]

# Paleta de cores
_C_HEADER    = "2E75B6"   # azul — cabeçalho de colunas
_C_TEAM      = "1F4E79"   # azul escuro — seção do time
_C_METRIC    = "D6E4F0"   # azul claro — rótulo de métrica
_C_ROW_ALT   = "EBF3FB"   # azul muito claro — linha alternada
_C_WHITE     = "FFFFFF"
_C_RESULTADO = {"Vitória": "C6EFCE", "Empate": "FFEB9C", "Derrota": "FFC7CE"}


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 28)


def _format_data_sheet(ws, col_labels: dict[str, str]):
    """Aplica formatação visual a uma aba de dados."""
    # Renomeia cabeçalhos para rótulos legíveis
    for cell in ws[1]:
        cell.value = col_labels.get(cell.value, cell.value)
        cell.font      = Font(bold=True, color=_C_WHITE)
        cell.fill      = _fill(_C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    ws.row_dimensions[1].height = 20

    # Linhas de dados
    res_col = None
    for cell in ws[1]:
        if cell.value == "resultado":
            res_col = cell.column

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_fill = _fill(_C_ROW_ALT) if i % 2 == 0 else _fill(_C_WHITE)
        resultado_val = None
        for cell in row:
            if cell.column == res_col:
                resultado_val = cell.value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()

        # Cor da linha pelo resultado
        resultado_fill = _fill(_C_RESULTADO.get(resultado_val, "")) if resultado_val else row_fill
        for cell in row:
            cell.fill = resultado_fill if resultado_val else row_fill

    ws.freeze_panes = "A2"
    _auto_width(ws)


_C_ADV_DARK  = "843C0C"   # laranja escuro — grupo adversário
_C_ADV_HDR   = "C55A11"   # laranja médio — cabeçalho colunas adversário
_C_ADV_ROW   = "FCE4D6"   # laranja claro — linhas alternadas adversário


def _append_resumo(ws, tdf: pd.DataFrame, num_cols: list[tuple], start_row: int):
    """Escreve a tabela de resumo abaixo dos dados na aba do time."""
    team_cols = [(col, lbl) for col, lbl in num_cols
                 if col in tdf.columns and not col.startswith("adv_")]
    adv_cols  = [(col, lbl) for col, lbl in num_cols
                 if col in tdf.columns and col.startswith("adv_")]
    # Rótulos curtos para colunas do adversário (sem prefixo "Adv.")
    adv_cols_short = [(col, lbl.replace("Adv. ", "")) for col, lbl in adv_cols]
    all_cols  = team_cols + adv_cols_short

    n_team = len(team_cols)
    n_adv  = len(adv_cols)
    n_total = n_team + n_adv
    row = start_row

    # ── Linha 1: título da seção ──────────────────────────────────────────
    ws.cell(row, 1).value     = "RESUMO ESTATÍSTICO"
    ws.cell(row, 1).font      = Font(bold=True, color=_C_WHITE, size=11)
    ws.cell(row, 1).fill      = _fill(_C_TEAM)
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_total + 1)
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Linha 2: grupos "DO TIME" / "DO ADVERSÁRIO" ───────────────────────
    ws.cell(row, 1).border = _thin_border()  # célula vazia "Métrica"
    if n_team:
        ws.cell(row, 2).value     = "DO TIME"
        ws.cell(row, 2).font      = Font(bold=True, color=_C_WHITE)
        ws.cell(row, 2).fill      = _fill(_C_TEAM)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 2).border    = _thin_border()
        if n_team > 1:
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row, end_column=n_team + 1)
    if n_adv:
        adv_start_col = n_team + 2
        ws.cell(row, adv_start_col).value     = "DO ADVERSÁRIO"
        ws.cell(row, adv_start_col).font      = Font(bold=True, color=_C_WHITE)
        ws.cell(row, adv_start_col).fill      = _fill(_C_ADV_DARK)
        ws.cell(row, adv_start_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, adv_start_col).border    = _thin_border()
        if n_adv > 1:
            ws.merge_cells(start_row=row, start_column=adv_start_col,
                           end_row=row, end_column=n_total + 1)
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Linha 3: cabeçalhos de cada coluna ───────────────────────────────
    ws.cell(row, 1).value     = "Métrica"
    ws.cell(row, 1).font      = Font(bold=True, color=_C_WHITE)
    ws.cell(row, 1).fill      = _fill(_C_HEADER)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border    = _thin_border()
    for j, (_, lbl) in enumerate(all_cols, 2):
        is_adv  = j >= n_team + 2
        c_color = _C_ADV_HDR if is_adv else _C_HEADER
        c           = ws.cell(row, j)
        c.value     = lbl
        c.font      = Font(bold=True, color=_C_WHITE)
        c.fill      = _fill(c_color)
        c.alignment = Alignment(horizontal="center")
        c.border    = _thin_border()
    row += 1

    # ── Linhas de métricas ────────────────────────────────────────────────
    metrics_start = row
    for m_idx, (metric_name, func) in enumerate(_METRICS):
        team_fill = _fill(_C_METRIC)  if m_idx % 2 == 0 else _fill(_C_WHITE)
        adv_fill  = _fill(_C_ADV_ROW) if m_idx % 2 == 0 else _fill(_C_WHITE)

        lbl_cell           = ws.cell(row, 1)
        lbl_cell.value     = metric_name
        lbl_cell.font      = Font(bold=True)
        lbl_cell.fill      = team_fill
        lbl_cell.alignment = Alignment(horizontal="left")
        lbl_cell.border    = _thin_border()

        for j, (col, _) in enumerate(all_cols, 2):
            series = pd.to_numeric(tdf[col], errors="coerce").dropna()
            if len(series) >= 2:
                val = round(float(func(series)), 1)
            elif len(series) == 1:
                val = round(float(series.iloc[0]), 1)
            else:
                val = None
            is_adv          = j >= n_team + 2
            c               = ws.cell(row, j)
            c.value         = val
            c.fill          = adv_fill if is_adv else team_fill
            c.alignment     = Alignment(horizontal="center")
            c.border        = _thin_border()
            c.number_format = "0.0"
        row += 1

    # ── Formatação condicional por coluna ─────────────────────────────────
    metrics_end = row - 1
    for j in range(2, n_total + 2):
        col_letter = get_column_letter(j)
        ws.conditional_formatting.add(
            f"{col_letter}{metrics_start}:{col_letter}{metrics_end}",
            ColorScaleRule(
                start_type="min",      start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max",        end_color="63BE7B",
            ),
        )


def export_to_excel(all_rows: list, filename: str = "sofascore_stats.xlsx"):
    if not all_rows:
        print("\nNenhum dado coletado para exportar.")
        return

    df = pd.DataFrame(all_rows)

    col_labels = {
        "time": "Time", "data": "Data", "competicao": "Competição",
        "adversario": "Adversário", "local": "Local", "placar": "Placar",
        "resultado": "Resultado",
        **{col: lbl for col, lbl in _NUM_COLS},
    }

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Aba Todos
        df.to_excel(writer, sheet_name="Todos", index=False)
        _format_data_sheet(writer.sheets["Todos"], col_labels)

        # Aba por time: dados + resumo na mesma aba
        for team in df["time"].unique():
            tdf        = df[df["time"] == team]
            sheet_name = str(team)[:31]
            tdf.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _format_data_sheet(ws, col_labels)
            # Resumo começa 2 linhas abaixo dos dados
            resumo_start = len(tdf) + 3
            _append_resumo(ws, tdf, _NUM_COLS, resumo_start)

    print(f"\nExportado para: {filename}")

# ---------------------------------------------------------------------------
# Todos os times da Copa do Mundo 2026 (IDs do SofaScore)
# ---------------------------------------------------------------------------
ALL_TEAMS: dict[int, str] = {
    4691:  "Argélia",
    4819:  "Argentina",
    4741:  "Austrália",
    4718:  "Áustria",
    4717:  "Bélgica",
    4479:  "Bósnia e Herzegovina",
    4748:  "Brasil",
    4753:  "Cabo Verde",
    4752:  "Canadá",
    4820:  "Colômbia",
    4715:  "Croácia",
    55827: "Curaçao",
    4714:  "República Tcheca",
    4768:  "Costa do Marfim",
    4823:  "Congo RD",
    4757:  "Equador",
    4758:  "Egito",
    4713:  "Inglaterra",
    4481:  "França",
    4711:  "Alemanha",
    4764:  "Gana",
    7229:  "Haiti",
    4766:  "Irã",
    4767:  "Iraque",
    4770:  "Japão",
    4771:  "Jordânia",
    4781:  "México",
    4778:  "Marrocos",
    4705:  "Países Baixos",
    4784:  "Nova Zelândia",
    4475:  "Noruega",
    5164:  "Panamá",
    4789:  "Paraguai",
    4704:  "Portugal",
    4792:  "Catar",
    4834:  "Arábia Saudita",
    4695:  "Escócia",
    4739:  "Senegal",
    4736:  "África do Sul",
    4735:  "Coreia do Sul",
    4698:  "Espanha",
    4688:  "Suécia",
    4699:  "Suíça",
    4729:  "Tunísia",
    4700:  "Turquia",
    4724:  "Estados Unidos",
    4725:  "Uruguai",
    4723:  "Uzbequistão",
}


# ---------------------------------------------------------------------------
# CLI interativo
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    """Minúsculas sem acentos — para comparação fuzzy."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def _find_team(query: str, pool: dict[int, str]) -> list[tuple[int, str]]:
    norm_q = _normalize(query)

    exact = [(tid, n) for tid, n in pool.items() if _normalize(n) == norm_q]
    if exact:
        return exact

    prefix = [(tid, n) for tid, n in pool.items() if _normalize(n).startswith(norm_q)]
    if prefix:
        return prefix

    close_norms = difflib.get_close_matches(norm_q, [_normalize(n) for n in pool.values()], n=3, cutoff=0.5)
    return [(tid, n) for tid, n in pool.items() if _normalize(n) in close_norms]


def select_teams() -> dict[int, str]:
    sorted_teams = sorted(ALL_TEAMS.items(), key=lambda x: x[1])
    remaining    = dict(sorted_teams)
    selected: dict[int, str] = {}

    print("\n" + "=" * 56)
    print("    COPA DO MUNDO 2026 — COLETOR DE ESTATÍSTICAS")
    print("=" * 56)
    print("\nTimes disponíveis:\n")
    col_w = max(len(n) for n in ALL_TEAMS.values()) + 2
    for i, (_, name) in enumerate(sorted_teams, 1):
        print(f"  {name:<{col_w}}", end="\n" if i % 3 == 0 else "")
    print("\n")
    print("  -> Digite parte do nome para adicionar um time.")
    print("  -> 'todos' para selecionar todos os 48 times.")
    print("  -> 'pronto' para finalizar a selecao.\n")

    while True:
        sel_txt = ", ".join(selected.values()) if selected else "nenhum"
        try:
            raw = input(f"[Selecionados: {sel_txt}]\n> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break

        if not raw:
            continue

        cmd = raw.lower()

        if cmd == "pronto":
            if not selected:
                print("  Selecione ao menos um time.\n")
                continue
            break

        if cmd == "todos":
            selected = dict(ALL_TEAMS)
            print(f"  Todos os {len(selected)} times adicionados.")
            break

        matches = _find_team(raw, remaining)

        if not matches:
            print(f"  Time não encontrado: '{raw}'\n")
            continue

        if len(matches) == 1:
            tid, name = matches[0]
            selected[tid] = name
            del remaining[tid]
            print(f"  + {name} adicionado.\n")
        else:
            print("  Mais de um resultado:")
            for i, (_, name) in enumerate(matches, 1):
                print(f"    {i}. {name}")
            try:
                choice = input("  Número (Enter para cancelar): ").strip()
                if choice.isdigit() and 1 <= int(choice) <= len(matches):
                    tid, name = matches[int(choice) - 1]
                    selected[tid] = name
                    del remaining[tid]
                    print(f"  + {name} adicionado.\n")
                else:
                    print()
            except (EOFError, KeyboardInterrupt):
                print()

    return selected


def _ask_int(prompt: str, default: int) -> int:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
        return int(raw) if raw else default
    except (ValueError, EOFError, KeyboardInterrupt):
        return default


def _ask_str(prompt: str, default: str) -> str:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
        return raw if raw else default
    except (EOFError, KeyboardInterrupt):
        return default


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    _load_cache()
    teams   = select_teams()
    n_games = _ask_int("\nQuantos jogos recentes por time?", default=5)
    outfile = _ask_str("Nome do arquivo Excel?",            default="copa2026_stats.xlsx")

    print()
    session = BrowserSession()
    try:
        all_rows = []
        for team_id, team_name in teams.items():
            rows = collect_team_stats(session, team_id, team_name, n_games=n_games)
            all_rows.extend(rows)
        export_to_excel(all_rows, filename=outfile)
    finally:
        session.close()
